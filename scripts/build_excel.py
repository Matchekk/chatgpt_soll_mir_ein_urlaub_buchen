from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils import (
    CANDIDATE_COLUMNS,
    CANDIDATES_PATH,
    EXCEL_PATH,
    LINK_INTAKE_COLUMNS,
    LINK_INTAKE_PATH,
    ensure_directories,
    parse_float,
    read_csv,
)

TITLE_FILL = PatternFill("solid", fgColor="16324F")
HEADER_FILL = PatternFill("solid", fgColor="2F6F73")
SUBTLE_FILL = PatternFill("solid", fgColor="EEF4F3")
CARD_FILL = PatternFill("solid", fgColor="F7FAF9")
RED_FILL = PatternFill("solid", fgColor="F8D7DA")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_FONT = Font(name="Aptos", size=10)
TITLE_FONT = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D7DEE2")
MEDIUM = Side(style="medium", color="9EB5B8")

SCORE_COLUMNS = [
    "private_level",
    "child_potential_0_10",
    "quiet_score_0_10",
    "review_evidence_0_10",
    "beach_fit_0_10",
    "transfer_score_0_10",
    "budget_score_0_10",
    "overall_score_0_10",
]

EURO_COLUMNS = ["price_total", "price_per_person", "price_per_person_per_night"]
KM_COLUMNS = ["alc_km_est", "vlc_km_est"]


def numeric(value):
    number = parse_float(value)
    return number if number is not None else value


def candidate_frame() -> pd.DataFrame:
    df = read_csv(CANDIDATES_PATH, CANDIDATE_COLUMNS)
    for col in CANDIDATE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[CANDIDATE_COLUMNS]


def intake_frame() -> pd.DataFrame:
    df = read_csv(LINK_INTAKE_PATH, LINK_INTAKE_COLUMNS)
    for col in LINK_INTAKE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[LINK_INTAKE_COLUMNS]


def as_display_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in SCORE_COLUMNS + EURO_COLUMNS + KM_COLUMNS + ["nights", "review_count", "rating", "alc_drive_min_est", "vlc_drive_min_est"]:
        if col in out.columns:
            out[col] = out[col].map(numeric)
    return out


def style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if not cell.font or cell.font == Font():
                cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical="top", wrap_text=cell.alignment.wrap_text)


def write_table(ws, df: pd.DataFrame, table_name: str, start_row: int = 1, start_col: int = 1) -> None:
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start_col):
        cell = ws.cell(start_row, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=MEDIUM)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, start_col):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = Border(bottom=THIN)
            if headers[col_idx - start_col] in {"notes", "family_red_flags", "quiet_evidence", "review_evidence", "exclusion_reason"}:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if headers[col_idx - start_col] == "url" and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"
    end_row = max(start_row + len(df), start_row + 1)
    end_col = start_col + len(headers) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = ws.cell(start_row + 1, start_col + 1)
    ws.auto_filter.ref = ref
    for idx, header in enumerate(headers, start_col):
        width = min(max(len(str(header)) + 3, 12), 38)
        if header in {"url", "raw_markdown_path", "raw_json_path", "screenshot_path"}:
            width = 34
        elif header in {"notes", "family_red_flags", "quiet_evidence", "review_evidence"}:
            width = 42
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_conditional_formatting(ws, headers: list[str], first_row: int, last_row: int) -> None:
    if last_row < first_row:
        return
    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    if "overall_score_0_10" in header_map:
        col = get_column_letter(header_map["overall_score_0_10"])
        ws.conditional_formatting.add(f"{col}{first_row}:{col}{last_row}", ColorScaleRule(start_type="num", start_value=0, start_color="E06666", mid_type="num", mid_value=5, mid_color="FCE5CD", end_type="num", end_value=10, end_color="6AA84F"))
    if "child_potential_0_10" in header_map:
        col = get_column_letter(header_map["child_potential_0_10"])
        ws.conditional_formatting.add(f"{col}{first_row}:{col}{last_row}", ColorScaleRule(start_type="num", start_value=0, start_color="6AA84F", mid_type="num", mid_value=5, mid_color="FCE5CD", end_type="num", end_value=10, end_color="E06666"))
    if "budget_under_500pp" in header_map:
        col = get_column_letter(header_map["budget_under_500pp"])
        ws.conditional_formatting.add(f"{col}{first_row}:{col}{last_row}", CellIsRule(operator="equal", formula=['"true"'], fill=GREEN_FILL))
        ws.conditional_formatting.add(f"{col}{first_row}:{col}{last_row}", CellIsRule(operator="equal", formula=['"false"'], fill=RED_FILL))
    if "excluded" in header_map:
        col = get_column_letter(header_map["excluded"])
        ws.conditional_formatting.add(f"{col}{first_row}:{col}{last_row}", CellIsRule(operator="equal", formula=['"true"'], fill=RED_FILL))


def format_numeric_columns(ws, headers: list[str], first_row: int, last_row: int) -> None:
    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    for name in EURO_COLUMNS:
        if name in header_map:
            col = get_column_letter(header_map[name])
            for cell in ws[f"{col}{first_row}:{col}{last_row}"]:
                cell[0].number_format = '€#,##0.00'
    for name in SCORE_COLUMNS:
        if name in header_map:
            col = get_column_letter(header_map[name])
            for cell in ws[f"{col}{first_row}:{col}{last_row}"]:
                cell[0].number_format = "0.0"
    for name in KM_COLUMNS:
        if name in header_map:
            col = get_column_letter(header_map[name])
            for cell in ws[f"{col}{first_row}:{col}{last_row}"]:
                cell[0].number_format = "0.0"


def top_rows(df: pd.DataFrame, sort_col: str, n: int = 5, ascending: bool = False, filter_expr=None) -> pd.DataFrame:
    work = df.copy()
    if filter_expr is not None:
        work = work[filter_expr(work)]
    work["_sort"] = work[sort_col].map(lambda v: parse_float(v) if parse_float(v) is not None else -999)
    work = work.sort_values("_sort", ascending=ascending).head(n)
    return work[["name", "location", "price_per_person", "overall_score_0_10", "child_potential_0_10", "quiet_score_0_10", "url"]]


def dashboard(ws, df: pd.DataFrame) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Costa Blanca Travel Matrix"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    total = len(df)
    excluded = int(df["excluded"].astype(str).str.lower().eq("true").sum()) if "excluded" in df else 0
    blocked = int(df["needs_manual_input"].astype(str).str.lower().eq("true").sum()) if "needs_manual_input" in df else 0
    under_budget = int(df["budget_under_500pp"].astype(str).str.lower().eq("true").sum()) if "budget_under_500pp" in df else 0
    metrics = [("Kandidaten", total), ("Unter Budget", under_budget), ("Ausgeschlossen", excluded), ("Manual Input", blocked)]
    for i, (label, value) in enumerate(metrics):
        col = 1 + i * 2
        ws.cell(3, col, label).fill = SUBTLE_FILL
        ws.cell(3, col, label).font = Font(bold=True, color="16324F")
        ws.cell(4, col, value).font = Font(size=16, bold=True, color="16324F")
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)

    constraints = [
        "2 Erwachsene, max. 500 EUR p.P. / 1000 EUR gesamt",
        "Reisefenster 04.09.2026 bis 18.09.2026, sinnvoll max. 9 Nächte",
        "Kinder-/Family-/Resort-Signale werden streng abgewertet",
        "ALC/VLC-Distanzen sind Koordinaten-Näherungen, keine Live-Routen",
    ]
    ws["A6"] = "Harte Constraints"
    ws["A6"].font = Font(bold=True, color="16324F")
    for idx, text in enumerate(constraints, 7):
        ws.cell(idx, 1, text)

    recommendation = "Noch keine belastbare Empfehlung."
    available = df[df["excluded"].astype(str).str.lower() != "true"].copy() if "excluded" in df else df.copy()
    if not available.empty:
        available["_score"] = available["overall_score_0_10"].map(lambda v: parse_float(v) or -1)
        best = available.sort_values("_score", ascending=False).iloc[0]
        recommendation = f"Aktuell vorne: {best.get('name', 'unknown')} ({best.get('location', 'unknown')}) mit Score {best.get('overall_score_0_10', 'unknown')}."
    ws["E6"] = "Kurze Empfehlung"
    ws["E6"].font = Font(bold=True, color="16324F")
    ws.merge_cells("E7:H9")
    ws["E7"] = recommendation
    ws["E7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["E7"].fill = CARD_FILL

    sections = [
        ("Top 5 Gesamtscore", top_rows(df, "overall_score_0_10")),
        ("Top 5 unter Budget", top_rows(df, "overall_score_0_10", filter_expr=lambda d: d["budget_under_500pp"].astype(str).str.lower().eq("true"))),
        ("Top 5 niedrigstes Kinderpotenzial", top_rows(df, "child_potential_0_10", ascending=True)),
        ("Top 5 Ruhe", top_rows(df, "quiet_score_0_10")),
    ]
    start_rows = [12, 21, 30, 39]
    for (title, table_df), start in zip(sections, start_rows):
        ws.cell(start, 1, title).font = Font(bold=True, color="16324F")
        ws.cell(start, 1).fill = SUBTLE_FILL
        display = as_display_df(table_df.fillna(""))
        write_simple_block(ws, display, start + 1, 1)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20


def write_simple_block(ws, df: pd.DataFrame, start_row: int, start_col: int) -> None:
    for c, header in enumerate(df.columns, start_col):
        cell = ws.cell(start_row, c, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    for r, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c, value in enumerate(row, start_col):
            cell = ws.cell(r, c, value)
            cell.border = Border(bottom=THIN)
            if df.columns[c - start_col] == "url" and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.style = "Hyperlink"


def scoring_explanation(ws) -> None:
    rows = [
        ["Komponente", "Gewicht", "Logik"],
        ["private_level", "18%", "Privatsphäre der Unterkunft: ganze Wohnung/Haus besser, Hotel/Room schlechter."],
        ["child_potential_inverse", "18%", "10 minus child_potential_0_10. Niedriges Kinderpotenzial ist gut."],
        ["quiet_score_0_10", "18%", "Lage, Unterkunftstyp, Review-Hinweise, Anlagen-/Promenadenrisiko."],
        ["beach_fit_0_10", "14%", "Gute Cala-/Costa-Blanca-Profile hoch, Promenaden-/Hotelstrände niedrig."],
        ["transfer_score_0_10", "10%", "ALC primär, VLC sekundär; Koordinaten-Näherung."],
        ["budget_score_0_10", "12%", "Budgetstaffel nach Preis pro Person."],
        ["review_evidence_0_10", "10%", "Review-Anzahl und konkrete Ruhe-/Lärm-Hinweise."],
        ["Excluded", "hart", "Bei excluded=true wird overall_score_0_10 auf 0 gesetzt."],
    ]
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL
    ws["C1"].fill = HEADER_FILL
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 92
    ws.freeze_panes = "A2"


def raw_status_frame(candidates: pd.DataFrame, intake: pd.DataFrame) -> pd.DataFrame:
    columns = ["candidate_id", "name", "url", "crawl_status", "needs_manual_input", "raw_markdown_path", "raw_json_path", "screenshot_path", "last_updated"]
    left = candidates[[col for col in columns if col in candidates.columns]].copy()
    intake_small = intake[["link_id", "source_url", "status", "crawl_status", "needs_manual_input", "last_updated"]].copy()
    intake_small = intake_small.rename(columns={"source_url": "url"})
    return pd.concat([left, intake_small], ignore_index=True).fillna("")


def build_excel() -> Path:
    ensure_directories()
    candidates = candidate_frame()
    intake = intake_frame()
    wb = Workbook()
    wb.remove(wb.active)

    ws_dashboard = wb.create_sheet("Dashboard")
    dashboard(ws_dashboard, candidates)

    ws_candidates = wb.create_sheet("Kandidaten")
    candidate_display = as_display_df(candidates)
    write_table(ws_candidates, candidate_display, "CandidatesTable")
    add_conditional_formatting(ws_candidates, list(candidate_display.columns), 2, len(candidate_display) + 1)
    format_numeric_columns(ws_candidates, list(candidate_display.columns), 2, len(candidate_display) + 1)

    ws_matrix = wb.create_sheet("Bewertungsmatrix")
    matrix_cols = [
        "name",
        "location",
        "private_level",
        "child_potential_0_10",
        "quiet_score_0_10",
        "beach_fit_0_10",
        "transfer_score_0_10",
        "budget_score_0_10",
        "review_evidence_0_10",
        "overall_score_0_10",
        "notes",
    ]
    matrix_df = as_display_df(candidates[[col for col in matrix_cols if col in candidates.columns]])
    write_table(ws_matrix, matrix_df, "ScoringMatrixTable")
    add_conditional_formatting(ws_matrix, list(matrix_df.columns), 2, len(matrix_df) + 1)
    format_numeric_columns(ws_matrix, list(matrix_df.columns), 2, len(matrix_df) + 1)

    ws_excluded = wb.create_sheet("Ausgeschlossen")
    excluded_df = candidates[candidates["excluded"].astype(str).str.lower().eq("true")] if "excluded" in candidates else candidates.iloc[0:0]
    write_table(ws_excluded, as_display_df(excluded_df), "ExcludedTable")
    add_conditional_formatting(ws_excluded, list(excluded_df.columns), 2, len(excluded_df) + 1)

    ws_flights = wb.create_sheet("Flüge")
    flight_rows = pd.DataFrame(
        [
            ["ALC", "Alicante Airport", "Primärer Flughafen", "Distanzen im Kandidatenblatt sind Näherungen."],
            ["VLC", "Valencia Airport", "Sekundärer Flughafen", "Keine Live-Google-Maps-Fahrzeiten."],
        ],
        columns=["Code", "Airport", "Rolle", "Notiz"],
    )
    write_table(ws_flights, flight_rows, "FlightsTable")

    ws_intake = wb.create_sheet("Link Intake")
    write_table(ws_intake, intake, "LinkIntakeTable")

    ws_scoring = wb.create_sheet("Scoring Erklärung")
    scoring_explanation(ws_scoring)

    ws_raw = wb.create_sheet("Raw Crawl Status")
    write_table(ws_raw, raw_status_frame(candidates, intake), "RawCrawlStatusTable")

    for ws in wb.worksheets:
        style_sheet(ws)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.row == 1 and ws.title not in {"Dashboard", "Scoring Erklärung"}:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_dashboard["A1"].font = TITLE_FONT
    ws_dashboard["A1"].alignment = Alignment(horizontal="center", vertical="center")

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    return EXCEL_PATH


def main() -> int:
    path = build_excel()
    print(f"Built {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
